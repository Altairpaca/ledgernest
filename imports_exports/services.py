"""导入导出服务：CSV/XLSX 解析、行级校验、导入执行、导出、备份。"""
import csv
import io
import json
import zipfile
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from django.core.files.uploadedfile import UploadedFile
from django.db import transaction as db_transaction

from transactions.models import (
    Account,
    Category,
    Transaction,
    TransactionSplit,
    TransactionType,
)

# 固定模板列（第一版不做字段映射器）
TEMPLATE_COLUMNS = ["日期", "类型", "金额", "账户", "目标账户", "分类", "交易对象", "描述", "标签"]

TYPE_ALIASES = {
    "expense": TransactionType.EXPENSE,
    "支出": TransactionType.EXPENSE,
    "收入": TransactionType.INCOME,
    "income": TransactionType.INCOME,
    "转账": TransactionType.TRANSFER,
    "transfer": TransactionType.TRANSFER,
    "调整": TransactionType.ADJUSTMENT,
    "adjustment": TransactionType.ADJUSTMENT,
    "退款": TransactionType.REFUND,
    "refund": TransactionType.REFUND,
    "报销": TransactionType.REIMBURSEMENT,
    "reimbursement": TransactionType.REIMBURSEMENT,
}


class ImportRowError(Exception):
    pass


def parse_amount(value) -> Decimal:
    if value in (None, ""):
        raise ImportRowError("金额为空")
    try:
        amt = Decimal(str(value).replace(",", "").strip())
    except InvalidOperation:
        raise ImportRowError(f"金额无法解析：{value}")
    if amt == 0:
        raise ImportRowError("金额为 0")
    return amt


def parse_date(value) -> date:
    if value in (None, ""):
        raise ImportRowError("日期为空")
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ImportRowError(f"日期无法解析：{value}（支持 YYYY-MM-DD）")


def read_upload(file: UploadedFile) -> list[dict]:
    """读取上传文件为行字典列表（保持原始字符串值）。"""
    name = (file.name or "").lower()
    raw = file.read()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            return []
        header = [str(h).strip() if h is not None else "" for h in header]
        out = []
        for row in rows:
            out.append(dict(zip(header, row)))
        return out
    # CSV：自动探测编码
    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def validate_rows(ledger, rows: list[dict]) -> list[dict]:
    """逐行校验并转换为可入库数据；返回带 error 字段的行列表。"""
    accounts = {a.name: a for a in Account.objects.filter(ledger=ledger, is_active=True)}
    all_accounts = {a.name: a for a in Account.all_objects.filter(ledger=ledger)}
    categories = {c.name: c for c in Category.all_objects.filter(ledger=ledger)}

    out = []
    for idx, raw in enumerate(rows, start=2):  # Excel 行号从 2 开始（含表头）
        row = {"row_number": idx, "error": None, "data": None, "raw": raw}
        try:
            txn_type = TYPE_ALIASES.get(str(raw.get("类型", "")).strip())
            if txn_type is None:
                raise ImportRowError("类型必须为 支出/收入/转账/调整")
            amount = parse_amount(raw.get("金额"))
            date_val = parse_date(raw.get("日期"))
            account_name = str(raw.get("账户", "")).strip()
            to_account_name = str(raw.get("目标账户", "")).strip()
            category_name = str(raw.get("分类", "")).strip()
            description = str(raw.get("描述", "") or "").strip()
            counterparty = str(raw.get("交易对象", "") or "").strip()
            tag_names = [t.strip() for t in str(raw.get("标签", "") or "").split(",") if t.strip()]

            account = accounts.get(account_name) or all_accounts.get(account_name)
            if txn_type in (
                TransactionType.EXPENSE, TransactionType.INCOME, TransactionType.ADJUSTMENT,
                TransactionType.REFUND, TransactionType.REIMBURSEMENT,
            ):
                if account is None:
                    raise ImportRowError(f"账户「{account_name}」不存在")
            if txn_type == TransactionType.TRANSFER:
                to_account = accounts.get(to_account_name) or all_accounts.get(to_account_name)
                if account is None or to_account is None:
                    raise ImportRowError("转账必须指定两个存在的账户（账户/目标账户）")
                if account.id == to_account.id:
                    raise ImportRowError("转账账户不能相同")
            else:
                to_account = None
                if to_account_name:
                    raise ImportRowError("非转账类型不能填写目标账户")

            category = None
            if category_name:
                category = categories.get(category_name)
                if category is None:
                    raise ImportRowError(f"分类「{category_name}」不存在")
                if category.kind == "income" and txn_type not in (TransactionType.INCOME, TransactionType.REFUND, TransactionType.REIMBURSEMENT):
                    raise ImportRowError(f"分类「{category_name}」是收入分类")
                if category.kind == "expense" and txn_type not in (TransactionType.EXPENSE, TransactionType.REFUND, TransactionType.REIMBURSEMENT):
                    raise ImportRowError(f"分类「{category_name}」是支出分类")

            if txn_type == TransactionType.ADJUSTMENT and amount < 0:
                amount_base = amount
                currency = "CNY"
                rate = Decimal("1")
            else:
                currency = str(raw.get("币种", "") or "").strip() or ledger.base_currency
                try:
                    rate = Decimal(str(raw.get("汇率", "1") or "1").strip())
                except InvalidOperation:
                    raise ImportRowError("汇率无法解析")
                amount_base = (amount * rate).quantize(Decimal("0.01"))

            row["data"] = {
                "type": txn_type,
                "date": date_val,
                "amount": abs(amount) if txn_type != TransactionType.ADJUSTMENT else amount,
                "currency": currency,
                "exchange_rate": rate,
                "amount_base": amount_base,
                "from_account": account,
                "to_account": to_account,
                "category": category,
                "counterparty": counterparty,
                "description": description,
                "tag_names": tag_names,
            }
        except ImportRowError as exc:
            row["error"] = str(exc)
        out.append(row)
    return out


def import_rows(ledger, validated_rows, actor=None, batch_label="") -> dict:
    """执行导入：仅导入有效行（无效行跳过并在结果中报告）。"""
    from audit.services import audit_log

    valid = [r for r in validated_rows if r["error"] is None and r["data"]]
    created = 0
    skipped_duplicates = 0
    with db_transaction.atomic():
        existing = _existing_fingerprints(ledger)
        for r in valid:
            fp = _fingerprint(r["data"])
            if fp in existing:
                skipped_duplicates += 1
                r["error"] = "与已有流水重复，已跳过"
                continue
            existing.add(fp)
            txn = Transaction.objects.create(
                ledger=ledger,
                type=r["data"]["type"],
                date=r["data"]["date"],
                amount=r["data"]["amount"],
                currency=r["data"]["currency"],
                exchange_rate=r["data"]["exchange_rate"],
                amount_base=r["data"]["amount_base"],
                from_account=r["data"]["from_account"],
                to_account=r["data"]["to_account"],
                category=r["data"]["category"],
                counterparty=r["data"]["counterparty"],
                description=r["data"]["description"],
                created_by=actor,
                updated_by=actor,
            )
            from transactions.models import Tag

            for name in r["data"]["tag_names"][:10]:
                tag, _ = Tag.objects.get_or_create(ledger=ledger, name=name)
                txn.tags.add(tag)
            created += 1
    audit_log(
        actor=actor, ledger=ledger, action="import", object_type="import",
        summary=f"导入流水：成功 {created}，失败 {len(validated_rows) - len(valid)}",
        changes={"total": len(validated_rows), "created": created, "errors": len(validated_rows) - len(valid)},
    )
    return {
        "total": len(validated_rows),
        "valid": len(valid),
        "created": created,
        "skipped_duplicates": skipped_duplicates,
        "errors": [r for r in validated_rows if r["error"]],
    }


def _fingerprint(data: dict) -> str:
    return "|".join(
        [
            str(data["date"]),
            data["type"],
            f"{data['amount']:.2f}",
            str(data["from_account"].id if data["from_account"] else ""),
            str(data["to_account"].id if data["to_account"] else ""),
            data["description"],
        ]
    )


def _existing_fingerprints(ledger):
    """近期同类流水的指纹集合，用于基础重复检测。"""
    recent = Transaction.objects.filter(ledger=ledger, date__gte=date.today().replace(year=date.today().year - 1))
    return {
        "|".join(
            [
                str(t.date),
                t.type,
                f"{t.amount:.2f}",
                str(t.from_account_id or ""),
                str(t.to_account_id or ""),
                t.description,
            ]
        )
        for t in recent.only("date", "type", "amount", "from_account_id", "to_account_id", "description")
    }


# ---------------------------------------------------------------------------
# 导出
# ---------------------------------------------------------------------------
def export_transactions_csv(ledger, qs) -> str:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(
        ["日期", "类型", "金额", "币种", "汇率", "基础币金额", "账户", "目标账户", "分类", "交易对象", "描述", "标签", "创建人"]
    )
    for t in qs.select_related("from_account", "to_account", "category", "created_by").prefetch_related("tags"):
        writer.writerow(
            [
                t.date.isoformat(),
                t.get_type_display(),
                t.amount,
                t.currency,
                t.exchange_rate,
                t.amount_base,
                t.from_account.name if t.from_account else "",
                t.to_account.name if t.to_account else "",
                t.category.name if t.category else "",
                t.counterparty,
                t.description,
                ",".join(t.tags.values_list("name", flat=True)),
                t.created_by.effective_display_name if t.created_by else "",
            ]
        )
    return buf.getvalue()


def export_transactions_xlsx(ledger, qs) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "流水"
    ws.append(TEMPLATE_COLUMNS + ["币种", "汇率", "基础币金额", "创建人"])
    for c in ws[1]:
        c.font = Font(bold=True)
    for t in qs.select_related("from_account", "to_account", "category", "created_by").prefetch_related("tags"):
        ws.append(
            [
                t.date.isoformat(),
                t.get_type_display(),
                float(t.amount),
                t.from_account.name if t.from_account else "",
                t.to_account.name if t.to_account else "",
                t.category.name if t.category else "",
                t.counterparty,
                t.description,
                ",".join(t.tags.values_list("name", flat=True)),
                t.currency,
                float(t.exchange_rate),
                float(t.amount_base),
                t.created_by.effective_display_name if t.created_by else "",
            ]
        )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_report_csv(rows: list[dict], columns: list[str]) -> str:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(columns)
    for r in rows:
        writer.writerow([r.get(c) if c in r else r.get("label", "") for c in columns])
    return buf.getvalue()


def export_report_xlsx(rows: list[dict], columns: list[str]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.append(columns)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append([r.get(c) if c in r else r.get("label", "") for c in columns])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# 备份与恢复
# ---------------------------------------------------------------------------
def backup_ledger(ledger) -> bytes:
    """账本结构化备份（不含密码与认证信息）。"""
    from accounts.models import User
    from ledgers.models import Ledger, LedgerMembership
    from reports.models import ReportDefinition
    from transactions.models import Account as TAccount
    from transactions.models import Budget, Category, Tag, Transaction, TransactionSplit

    def ser_date(d):
        return d.isoformat() if d else None

    def ser_dt(dt):
        return dt.isoformat() if dt else None

    data = {
        "version": 1,
        "exported_at": datetime.now().isoformat(),
        "ledger": {
            "name": ledger.name,
            "description": ledger.description,
            "base_currency": ledger.base_currency,
            "timezone": ledger.timezone,
            "fiscal_year_start_month": ledger.fiscal_year_start_month,
            "archived_at": ser_dt(ledger.archived_at),
        },
        "members": [
            {
                "username": m.user.username,
                "display_name": m.user.display_name,
                "email": m.user.email,
                "role": m.role,
                "joined_at": ser_dt(m.joined_at),
            }
            for m in LedgerMembership.objects.filter(ledger=ledger, is_active=True).select_related("user")
        ],
        "accounts": [
            {
                "name": a.name, "account_type": a.account_type, "currency": a.currency,
                "opening_balance": str(a.opening_balance), "is_active": a.is_active, "sort_order": a.sort_order,
            }
            for a in TAccount.all_objects.filter(ledger=ledger)
        ],
        "categories": [
            {
                "name": c.name, "kind": c.kind, "parent": c.parent.name if c.parent else None,
                "icon": c.icon, "sort_order": c.sort_order, "is_active": c.is_active,
            }
            for c in Category.all_objects.filter(ledger=ledger)
        ],
        "tags": [{"name": t.name, "color": t.color} for t in Tag.objects.filter(ledger=ledger)],
        "transactions": [
            {
                "type": t.type, "date": ser_date(t.date), "amount": str(t.amount),
                "currency": t.currency, "exchange_rate": str(t.exchange_rate),
                "amount_base": str(t.amount_base),
                "from_account": t.from_account.name if t.from_account else None,
                "to_account": t.to_account.name if t.to_account else None,
                "category": t.category.name if t.category else None,
                "counterparty": t.counterparty, "description": t.description,
                "created_by": t.created_by.username if t.created_by else None,
                "tags": list(t.tags.values_list("name", flat=True)),
                "splits": [
                    {
                        "category": s.category.name if s.category else None,
                        "amount": str(s.amount),
                    }
                    for s in TransactionSplit.objects.filter(transaction=t)
                ],
            }
            for t in Transaction.all_objects.filter(ledger=ledger).prefetch_related("tags")
        ],
        "budgets": [
            {
                "budget_type": b.budget_type, "category": b.category.name if b.category else None,
                "year": b.year, "month": b.month, "amount": str(b.amount),
            }
            for b in Budget.objects.filter(ledger=ledger)
        ],
        "report_definitions": [
            {
                "name": r.name, "description": r.description, "definition_json": r.definition_json,
                "is_shared": r.is_shared,
            }
            for r in ReportDefinition.objects.filter(ledger=ledger)
        ],
    }
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("ledger.json", json.dumps(data, ensure_ascii=False, indent=2))
        zf.writestr("README.txt", "LedgerNest 账本备份。恢复方式见 README。不包含密码与认证密钥。")
    return buf.getvalue()


def restore_ledger(backup_bytes: bytes, owner) -> dict:
    """从备份 zip 恢复账本（新账本，不覆盖现有数据）。"""
    from ledgers.models import Ledger, LedgerMembership
    from reports.models import ReportDefinition
    from transactions.models import Account as TAccount
    from transactions.models import Budget, Category, Tag, Transaction, TransactionSplit

    zf = zipfile.ZipFile(io.BytesIO(backup_bytes))
    data = json.loads(zf.read("ledger.json"))
    if data.get("version") != 1:
        raise ValueError("不支持的备份版本")

    ledger_info = data["ledger"]
    ledger = Ledger.objects.create(
        name=f"{ledger_info['name']}（恢复）",
        description=ledger_info.get("description", ""),
        owner=owner,
        base_currency=ledger_info.get("base_currency", "CNY"),
        timezone=ledger_info.get("timezone", "UTC"),
        fiscal_year_start_month=ledger_info.get("fiscal_year_start_month", 1),
    )
    LedgerMembership.objects.create(ledger=ledger, user=owner, role=10, invited_by=owner)

    account_map = {}
    for a in data.get("accounts", []):
        acc = TAccount.objects.create(
            ledger=ledger, name=a["name"], account_type=a.get("account_type", "bank"),
            currency=a.get("currency", "CNY"), opening_balance=Decimal(a.get("opening_balance", "0")),
            is_active=a.get("is_active", True), sort_order=a.get("sort_order", 0),
        )
        account_map[a["name"]] = acc

    category_map = {}
    for c in data.get("categories", []):
        cat = Category.objects.create(
            ledger=ledger, name=c["name"], kind=c.get("kind", "expense"),
            icon=c.get("icon", ""), sort_order=c.get("sort_order", 0),
            is_active=c.get("is_active", True),
        )
        category_map[(c["kind"], c["name"])] = cat
    for c in data.get("categories", []):
        if c.get("parent"):
            cat = category_map.get((c["kind"], c["name"]))
            parent = category_map.get((c["kind"], c["parent"]))
            if cat and parent:
                cat.parent = parent
                cat.save(update_fields=["parent"])

    tag_map = {}
    for t in data.get("tags", []):
        tag, _ = Tag.objects.get_or_create(ledger=ledger, name=t["name"], defaults={"color": t.get("color", "")})
        tag_map[t["name"]] = tag

    for t in data.get("transactions", []):
        txn = Transaction.objects.create(
            ledger=ledger, type=t["type"],
            date=date.fromisoformat(t["date"]),
            amount=Decimal(t["amount"]), currency=t.get("currency", "CNY"),
            exchange_rate=Decimal(t.get("exchange_rate", "1")),
            amount_base=Decimal(t.get("amount_base", t["amount"])),
            from_account=account_map.get(t.get("from_account")),
            to_account=account_map.get(t.get("to_account")),
            category=category_map.get(("expense", t["category"])) or category_map.get(("income", t["category"])),
            counterparty=t.get("counterparty", ""), description=t.get("description", ""),
            created_by=owner, updated_by=owner,
        )
        for tag_name in t.get("tags", []):
            if tag_name in tag_map:
                txn.tags.add(tag_map[tag_name])
        for s in t.get("splits", []):
            TransactionSplit.objects.create(
                transaction=txn,
                category=category_map.get(("expense", s["category"])) or category_map.get(("income", s["category"])),
                amount=Decimal(s["amount"]),
            )

    for b in data.get("budgets", []):
        Budget.objects.create(
            ledger=ledger, budget_type=b["budget_type"],
            category=category_map.get(("expense", b.get("category"))),
            year=b["year"], month=b["month"], amount=Decimal(b["amount"]),
        )

    for r in data.get("report_definitions", []):
        ReportDefinition.objects.create(
            ledger=ledger, name=r["name"], description=r.get("description", ""),
            definition_json=r.get("definition_json", {}), created_by=owner,
            is_shared=r.get("is_shared", True),
        )
    return {"ledger": ledger, "transactions": len(data.get("transactions", [])), "accounts": len(account_map)}
