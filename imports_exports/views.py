"""导入导出视图：模板下载、上传预览、确认导入、CSV/XLSX 导出、备份。"""
import io

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import redirect, render
from django.views.decorators.http import require_POST

from audit.services import audit_log
from core.models import EDIT_ROLES
from ledgers.views import _ensure_member

from .services import (
    TEMPLATE_COLUMNS,
    backup_ledger,
    export_report_csv,
    export_report_xlsx,
    export_transactions_csv,
    export_transactions_xlsx,
    import_rows,
    read_upload,
    validate_rows,
)


def _require_edit(request):
    membership = getattr(request, "membership", None)
    if membership is None or membership.role not in EDIT_ROLES:
        return HttpResponseForbidden("只读成员不能导入数据。")


@login_required
@_ensure_member
def import_view(request, ledger_pk):
    ledger = request.ledger
    denied = _require_edit(request)
    if denied:
        return denied
    return render(request, "imports/import.html", {"ledger": ledger, "template_columns": TEMPLATE_COLUMNS})


@login_required
@_ensure_member
@require_POST
def import_preview(request, ledger_pk):
    ledger = request.ledger
    denied = _require_edit(request)
    if denied:
        return denied
    file = request.FILES.get("file")
    if file is None:
        messages.error(request, "请选择要导入的文件。")
        return redirect("imports:import", ledger_pk=ledger_pk)
    if file.size > settings.DATA_UPLOAD_MAX_MEMORY_SIZE:
        messages.error(request, "文件过大（最大 20MB）。")
        return redirect("imports:import", ledger_pk=ledger_pk)
    if not (file.name.endswith(".csv") or file.name.endswith(".xlsx")):
        messages.error(request, "仅支持 CSV 或 XLSX 文件。")
        return redirect("imports:import", ledger_pk=ledger_pk)
    try:
        rows = read_upload(file)
    except Exception as exc:
        messages.error(request, f"文件解析失败：{exc}")
        return redirect("imports:import", ledger_pk=ledger_pk)
    if not rows:
        messages.error(request, "文件中没有数据行。")
        return redirect("imports:import", ledger_pk=ledger_pk)
    validated = validate_rows(ledger, rows)
    valid_count = sum(1 for r in validated if r["error"] is None)
    request.session["import_preview"] = {
        "rows": validated[:200],
        "total": len(validated),
        "valid_count": valid_count,
    }
    return render(
        request,
        "imports/preview.html",
        {
            "ledger": ledger,
            "preview": validated[:200],
            "total": len(validated),
            "valid_count": valid_count,
            "error_count": len(validated) - valid_count,
        },
    )


@login_required
@_ensure_member
@require_POST
def import_confirm(request, ledger_pk):
    """确认导入：仅导入有效行，单行错误不影响其他行。"""
    ledger = request.ledger
    denied = _require_edit(request)
    if denied:
        return denied
    preview = request.session.pop("import_preview", None)
    if preview is None:
        messages.error(request, "导入预览已失效，请重新上传。")
        return redirect("imports:import", ledger_pk=ledger_pk)
    result = import_rows(ledger, preview["rows"], actor=request.user, batch_label=request.POST.get("batch", ""))
    return render(
        request,
        "imports/result.html",
        {
            "ledger": ledger,
            "result": result,
            "errors": result["errors"][:50],
        },
    )


@login_required
@_ensure_member
def import_template(request, ledger_pk):
    """下载导入模板（XLSX）。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "流水模板"
    ws.append(TEMPLATE_COLUMNS)
    header_fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = header_fill
    ws.append(["2026-08-01", "支出", 25.5, "现金", "", "餐饮", "便利店", "买水", "日常"])
    ws.append(["2026-08-02", "收入", 8000, "工资卡", "", "工资", "公司", "八月工资", ""])
    ws.append(["2026-08-03", "转账", 1000, "工资卡", "支付宝", "", "", "转到支付宝", ""])
    ws.auto_filter.ref = ws.dimensions
    for col, width in zip("ABCDEFGHI", (14, 10, 12, 14, 12, 12, 14, 24, 16)):
        ws.column_dimensions[col].width = width
    buf = io.BytesIO()
    wb.save(buf)
    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="ledgernest_template.xlsx"'
    return response


# ---------------------------------------------------------------------------
# 导出
# ---------------------------------------------------------------------------
@login_required
@_ensure_member
def export_view(request, ledger_pk):
    return render(request, "imports/export.html", {"ledger": request.ledger})


@login_required
@_ensure_member
def export_transactions(request, ledger_pk):
    """导出当前筛选条件下的流水（CSV 或 XLSX）。"""
    from transactions.forms import TransactionFilterForm
    from transactions.models import Transaction

    fmt = request.GET.get("format", "csv")
    if fmt not in ("csv", "xlsx"):
        return HttpResponseForbidden("不支持的格式。")
    filter_form = TransactionFilterForm(request.ledger, request.GET or None)
    qs = Transaction.objects.filter(ledger=request.ledger)
    if filter_form.is_valid():
        qs = filter_form.apply(qs)
    qs = qs[: settings.EXPORT_ROW_LIMIT]

    if fmt == "csv":
        content = export_transactions_csv(request.ledger, qs)
        response = HttpResponse(content, content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="transactions.csv"'
    else:
        content = export_transactions_xlsx(request.ledger, qs)
        response = HttpResponse(
            content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="transactions.xlsx"'
    audit_log(
        actor=request.user, ledger=request.ledger, action="export", object_type="export",
        summary=f"导出流水（{fmt.upper()}）",
    )
    return response


@login_required
@_ensure_member
def export_report(request, ledger_pk):
    """导出报表结果（CSV/XLSX）。report_key 或 report_pk 二选一。"""
    from reports.services import (
        builtin_account,
        builtin_budget,
        builtin_cashflow,
        builtin_category,
        builtin_counterparty,
        builtin_member,
        builtin_tag,
        builtin_trend,
        run_custom_report,
    )

    fmt = request.GET.get("format", "csv")
    if fmt not in ("csv", "xlsx"):
        return HttpResponseForbidden("不支持的格式。")
    ledger = request.ledger
    rows, columns = [], ["label", "income", "expense", "value"]

    report_key = request.GET.get("report_key")
    if report_key:
        from datetime import date, timedelta

        start = date.fromisoformat(request.GET.get("start", (date.today() - timedelta(days=90)).isoformat()))
        end = date.fromisoformat(request.GET.get("end", date.today().isoformat()))
        if report_key == "trend":
            result = builtin_trend(ledger, months=int(request.GET.get("months", 6)))
            rows = [
                {"label": r["label"], "income": r["income"], "expense": r["expense"], "value": r["net"]}
                for r in result["rows"]
            ]
            columns = ["月份", "收入", "支出", "净额"]
        elif report_key == "category":
            result = builtin_category(ledger, request.GET.get("kind", "expense"), start, end)
            rows = [{"label": r["name"], "income": "", "expense": "", "value": r["total"]} for r in result["items"]]
            columns = ["分类", "", "", "金额"]
        elif report_key == "account":
            result = builtin_account(ledger, start, end)
            rows = [
                {
                    "label": r["account"].name, "income": r["inflow"], "expense": r["outflow"],
                    "value": r["balance"],
                }
                for r in result["rows"]
            ]
            columns = ["账户", "流入", "流出", "期末余额"]
        elif report_key == "member":
            result = builtin_member(ledger, start, end)
            rows = [
                {
                    "label": r["user"].effective_display_name if r["user"] else "未知",
                    "income": r["income"], "expense": r["expense"], "value": r["count"],
                }
                for r in result["rows"]
            ]
            columns = ["成员", "收入", "支出", "笔数"]
        elif report_key == "cashflow":
            result = builtin_cashflow(ledger, start, end)
            rows = [
                {"label": r["label"], "income": r["income"], "expense": r["expense"], "value": r["net"]}
                for r in result["rows"]
            ]
            columns = ["月份", "收入", "支出", "净额"]
        elif report_key == "tag":
            result = builtin_tag(ledger, start, end)
            rows = [
                {
                    "label": r["tag"].name, "income": r["income"], "expense": r["expense"], "value": r["count"],
                }
                for r in result["rows"]
            ]
            columns = ["标签", "收入", "支出", "次数"]
        elif report_key == "counterparty":
            result = builtin_counterparty(ledger, start, end)
            rows = [
                {
                    "label": r["counterparty"], "income": r["income"], "expense": r["expense"],
                    "value": r["count"],
                }
                for r in result["rows"]
            ]
            columns = ["交易对象", "收入", "支出", "次数"]
        elif report_key == "budget":
            result = builtin_budget(ledger, int(request.GET.get("year", start.year)), int(request.GET.get("month", start.month)))
            rows = [
                {"label": r["label"], "income": r["spent"], "expense": r["amount"], "value": r["percent"]}
                for r in result["rows"]
            ]
            columns = ["项目", "已支出", "预算", "执行%"]
    else:
        from reports.models import ReportDefinition

        report_pk = request.GET.get("report_pk")
        report = ReportDefinition.objects.get(pk=report_pk, ledger=ledger)
        result = run_custom_report(ledger, report.definition_json)
        rows = [
            {"label": r["label"], "income": r["income"], "expense": r["expense"], "value": r["value"]}
            for r in result["rows"]
        ]
        columns = ["维度", "收入", "支出", "指标值"]

    audit_log(
        actor=request.user, ledger=ledger, action="export", object_type="export",
        summary=f"导出报表（{fmt.upper()}）",
    )
    if fmt == "csv":
        response = HttpResponse(
            export_report_csv(rows, columns), content_type="text/csv; charset=utf-8"
        )
        response["Content-Disposition"] = 'attachment; filename="report.csv"'
    else:
        response = HttpResponse(
            export_report_xlsx(rows, columns),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="report.xlsx"'
    return response


# ---------------------------------------------------------------------------
# 备份
# ---------------------------------------------------------------------------
@login_required
@_ensure_member
def backup_download(request, ledger_pk):
    content = backup_ledger(request.ledger)
    response = HttpResponse(content, content_type="application/zip")
    response["Content-Disposition"] = f'attachment; filename="ledgernest-{request.ledger.pk}-backup.zip"'
    audit_log(
        actor=request.user, ledger=request.ledger, action="export", object_type="backup",
        summary="下载账本备份",
    )
    return response
